#!/usr/bin/env python3
"""
Genera un Excel con el listado de URLs de recursos (imagenes y PDF) a partir
de enlaces-recursos.json, para el pedido de migracion de enlaces rotos.

- Solo incluye imagenes y PDF (excluye fuentes .woff y .js).
- Una fila por (experimento, URL) para poder filtrar por campana activa.
- Columna "Experimento" = carpeta de primer nivel dentro de bitacora/.
"""

import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_IN = os.path.join(BASE_DIR, "enlaces-recursos.json")
XLSX_OUT = os.path.join(BASE_DIR, "URLs-recursos-campanas.xlsx")

# Extensiones incluidas: imagenes, PDF y fuentes
TIPOS_OK = {"png", "jpg", "jpeg", "gif", "webp", "svg", "pdf", "woff", "woff2"}


def tipo_legible(ext: str) -> str:
    if ext == "pdf":
        return "PDF"
    if ext in ("woff", "woff2"):
        return "Fuente"
    return "Imagen"


def main():
    with open(JSON_IN, encoding="utf-8") as fh:
        data = json.load(fh)

    # Construir filas: UNA por URL unica (sin duplicados)
    filas = []
    for r in data["recursos"]:
        if r["extension"] not in TIPOS_OK:
            continue
        experimentos = sorted({a.split(os.sep)[0] for a in r["archivos"]})
        filas.append({
            "experimento": ", ".join(experimentos),
            "tipo": tipo_legible(r["extension"]),
            "formato": r["extension"].upper(),
            "url": r["url"],
        })
    # Orden: tipo, url
    filas.sort(key=lambda x: (x["tipo"], x["url"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Recursos"

    # --- Encabezados ---
    cols = ["N°", "Usado en (experimentos)", "Tipo", "Formato", "URL del recurso"]
    ws.append(cols)

    azul = PatternFill("solid", fgColor="1F4E78")
    blanco_bold = Font(color="FFFFFF", bold=True, size=11)
    borde = Border(*[Side(style="thin", color="D9D9D9")] * 4)
    centro = Alignment(horizontal="center", vertical="center")
    izq = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for c, _ in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = azul
        cell.font = blanco_bold
        cell.alignment = centro
        cell.border = borde

    # --- Filas ---
    gris = PatternFill("solid", fgColor="F2F2F2")
    for i, f in enumerate(filas, 1):
        fila = i + 1
        valores = [i, f["experimento"], f["tipo"], f["formato"], f["url"]]
        for c, v in enumerate(valores, 1):
            cell = ws.cell(row=fila, column=c, value=v)
            cell.border = borde
            cell.alignment = centro if c in (1, 3, 4) else izq
            if i % 2 == 0:
                cell.fill = gris
        # URL como hipervinculo
        ws.cell(row=fila, column=5).hyperlink = f["url"]
        ws.cell(row=fila, column=5).font = Font(color="0563C1", underline="single")

    # --- Ancho de columnas ---
    anchos = {1: 6, 2: 34, 3: 10, 4: 10, 5: 120}
    for c, w in anchos.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # Congelar encabezado y activar autofiltro
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{len(filas) + 1}"

    wb.save(XLSX_OUT)

    # Resumen en consola
    total = len(filas)
    urls_unicas = len({f["url"] for f in filas})
    print(f"Filas (experimento x url): {total}")
    print(f"URLs unicas (imagen/PDF):  {urls_unicas}")
    print(f"Excel generado en: {XLSX_OUT}")


if __name__ == "__main__":
    main()
